from flask import Blueprint, request, jsonify
from datetime import datetime, timedelta
from openpyxl import load_workbook
import os
from werkzeug.utils import secure_filename

from models import db, WorkPlan, WorkPlanItem, Project, Material, RequiredMaterial
from auth import token_required, role_required
from project_access import require_project_access

workplan_bp = Blueprint('workplan_bp', __name__)

UPLOAD_FOLDER = '/app/uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


    project = db.get_or_404(Project, project_id)
    
    if project.status != 'active':
        return jsonify({'message': 'Нельзя составить план для неактивного проекта'}), 403
    
    existing_plan = WorkPlan.query.filter_by(project_id=project_id).first()
    if existing_plan:
        return jsonify({'message': 'План работ для этого проекта уже существует'}), 409
    
    data = request.get_json()
    if not data or not data.get('start_date') or not data.get('end_date') or not data.get('items'):
        return jsonify({'message': 'Требуются start_date, end_date и items'}), 400
    
    if len(data['items']) == 0:
        return jsonify({'message': 'План работ должен содержать хотя бы одну работу'}), 400
    
    try:
        start_date = datetime.fromisoformat(data['start_date'].replace('Z', '+00:00')).date()
        end_date = datetime.fromisoformat(data['end_date'].replace('Z', '+00:00')).date()
    except ValueError:
        return jsonify({'message': 'Неверный формат даты'}), 400
    
    if start_date >= end_date:
        return jsonify({'message': 'Дата начала должна быть раньше даты окончания'}), 400
    
    work_plan = WorkPlan(
        project_id=project_id,
        start_date=start_date,
        end_date=end_date
    )
    db.session.add(work_plan)
    db.session.flush()
    
    for index, item_data in enumerate(data['items']):
        if not item_data.get('name') or not item_data.get('quantity') or not item_data.get('unit'):
            return jsonify({'message': f'Элемент {index + 1}: требуются name, quantity и unit'}), 400
        
        if not item_data.get('start_date') or not item_data.get('end_date'):
            return jsonify({'message': f'Элемент {index + 1}: требуются start_date и end_date'}), 400
        
        try:
            item_start = datetime.fromisoformat(item_data['start_date'].replace('Z', '+00:00')).date()
            item_end = datetime.fromisoformat(item_data['end_date'].replace('Z', '+00:00')).date()
        except ValueError:
            return jsonify({'message': f'Элемент {index + 1}: неверный формат даты'}), 400
        
        if item_start >= item_end:
            return jsonify({'message': f'Элемент {index + 1}: дата начала должна быть раньше даты окончания'}), 400
        
        if item_start < start_date or item_end > end_date:
            return jsonify({'message': f'Элемент {index + 1}: сроки работы должны быть в пределах общего плана'}), 400
        
        work_plan_item = WorkPlanItem(
            work_plan_id=work_plan.id,
            name=item_data['name'],
            quantity=float(item_data['quantity']),
            unit=item_data['unit'],
            start_date=item_start,
            end_date=item_end,
            order=index
        )
        db.session.add(work_plan_item)
    
    db.session.commit()
    
    return jsonify(work_plan.to_dict()), 201


@workplan_bp.route('/api/projects/<int:project_id>/work-plan', methods=['GET'])
@token_required
def get_work_plan(project_id):
    """Получает план работ для проекта."""
    current_user = request.current_user
    
    access_error = require_project_access(project_id, current_user['id'], current_user['role'])
    if access_error:
        return access_error
    
    work_plan = WorkPlan.query.filter_by(project_id=project_id).first()
    if not work_plan:
        return jsonify({'message': 'План работ не найден'}), 404
    
    return jsonify(work_plan.to_dict()), 200


@workplan_bp.route('/api/projects/<int:project_id>/work-plan', methods=['PUT'])
@token_required
@role_required('client')
def update_work_plan(project_id):
    """Обновляет план работ для проекта, обрабатывая изменения в элементах."""
    work_plan = WorkPlan.query.filter_by(project_id=project_id).first()
    if not work_plan:
        return jsonify({'message': 'План работ не найден'}), 404

    data = request.get_json()
    if not data or 'items' not in data:
        return jsonify({'message': 'Отсутствуют данные для обновления'}), 400

    try:
        if 'start_date' in data and 'end_date' in data:
            work_plan.start_date = datetime.fromisoformat(data['start_date'].replace('Z', '+00:00')).date()
            work_plan.end_date = datetime.fromisoformat(data['end_date'].replace('Z', '+00:00')).date()

        existing_items = {item.id: item for item in work_plan.items}
        incoming_item_ids = {item['id'] for item in data['items'] if 'id' in item}

        for item_id, item in existing_items.items():
            if item_id not in incoming_item_ids:
                db.session.delete(item)

        for index, item_data in enumerate(data['items']):
            item_start = datetime.fromisoformat(item_data['start_date'].replace('Z', '+00:00')).date()
            item_end = datetime.fromisoformat(item_data['end_date'].replace('Z', '+00:00')).date()

            if item_start >= item_end:
                db.session.rollback()
                return jsonify({'message': f'Элемент \'{item_data["name"]}\': дата начала должна быть раньше даты окончания'}), 400

            if 'id' in item_data: 
                item = existing_items.get(item_data['id'])
                if item:
                    item.name = item_data['name']
                    item.quantity = float(item_data['quantity'])
                    item.unit = item_data['unit']
                    item.start_date = item_start
                    item.end_date = item_end
                    item.order = index
            else: 
                new_item = WorkPlanItem(
                    work_plan_id=work_plan.id,
                    name=item_data['name'],
                    quantity=float(item_data['quantity']),
                    unit=item_data['unit'],
                    start_date=item_start,
                    end_date=item_end,
                    order=index
                )
                db.session.add(new_item)

        work_plan.editing_status = 'edited'
        db.session.commit()

    except (ValueError, KeyError) as e:
        db.session.rollback()
        return jsonify({'message': f'Ошибка в данных: {str(e)}'}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Внутренняя ошибка сервера: {str(e)}'}), 500

    return jsonify(work_plan.to_dict()), 200


@workplan_bp.route('/api/projects/<int:project_id>/work-plan', methods=['DELETE'])
@token_required
@role_required('client')
def delete_work_plan(project_id):
    """Удаляет план работ для проекта."""
    current_user = request.current_user
    
    access_error = require_project_access(project_id, current_user['id'], current_user['role'])
    if access_error:
        return access_error
    
    work_plan = WorkPlan.query.filter_by(project_id=project_id).first()
    if not work_plan:
        return jsonify({'message': 'План работ не найден'}), 404
    
    db.session.delete(work_plan)
    db.session.commit()
    
    return jsonify({'message': 'План работ успешно удален'}), 200


@workplan_bp.route('/api/projects/<int:project_id>/work-plan/import', methods=['POST'])
@token_required
@role_required('client')
def import_work_plan(project_id):
    """Импортирует план работ из xlsx файла."""
    current_user = request.current_user
    
    access_error = require_project_access(project_id, current_user['id'], current_user['role'])
    if access_error:
        return access_error
    
    project = db.get_or_404(Project, project_id)
    
    existing_plan = WorkPlan.query.filter_by(project_id=project_id).first()
    if existing_plan:
        return jsonify({'message': 'План работ для этого проекта уже существует'}), 409
    
    if 'file' not in request.files:
        return jsonify({'message': 'Файл не найден'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'message': 'Файл не выбран'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'message': 'Недопустимый формат файла. Разрешены только xlsx и xls'}), 400
    
    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active
        
        start_date = None
        end_date = None
        items = []
        
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if not row[0]:
                continue
            
            first_cell = str(row[0]).strip()
            
            if 'срок' in first_cell.lower() and row[1]:
                date_str = str(row[1]).strip()
                if '-' in date_str:
                    dates = date_str.split('-')
                    if len(dates) == 2:
                        try:
                            start_date = datetime.strptime(dates[0].strip(), '%d.%m.%Y').date()
                            end_date = datetime.strptime(dates[1].strip(), '%d.%m.%Y').date()
                        except:
                            pass
            
            if row_idx > 5 and row[1] and row[2] and row[3]:
                try:
                    name = str(row[1]).strip()
                    unit = str(row[2]).strip()
                    quantity = float(row[3])
                    
                    if name and unit and quantity > 0 and len(name) > 3:
                        items.append({
                            'name': name,
                            'unit': unit,
                            'quantity': quantity
                        })
                except (ValueError, TypeError):
                    continue
        
        if not start_date or not end_date:
            return jsonify({'message': 'Не удалось определить сроки выполнения работ из файла'}), 400
        
        if len(items) == 0:
            return jsonify({'message': 'Не удалось извлечь работы из файла'}), 400
        
        work_plan = WorkPlan(
            project_id=project_id,
            start_date=start_date,
            end_date=end_date
        )
        db.session.add(work_plan)
        db.session.flush()
        
        total_days = (end_date - start_date).days
        days_per_item = max(1, total_days // len(items))
        
        current_start = start_date
        for index, item_data in enumerate(items):
            item_end = min(current_start + timedelta(days=days_per_item), end_date)
            if index == len(items) - 1:
                item_end = end_date
            
            work_plan_item = WorkPlanItem(
                work_plan_id=work_plan.id,
                name=item_data['name'],
                quantity=item_data['quantity'],
                unit=item_data['unit'],
                start_date=current_start,
                end_date=item_end,
                order=index
            )
            db.session.add(work_plan_item)
            current_start = item_end
        
        db.session.commit()
        
        return jsonify(work_plan.to_dict()), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Ошибка при импорте файла: {str(e)}'}), 500


@workplan_bp.route('/api/work-plan-items/<int:item_id>', methods=['GET'])
@token_required
def get_work_plan_item(item_id):
    """Получает конкретную работу плана."""
    current_user = request.current_user
    
    item = db.get_or_404(WorkPlanItem, item_id)
    work_plan = db.get_or_404(WorkPlan, item.work_plan_id)
    
    access_error = require_project_access(work_plan.project_id, current_user['id'], current_user['role'])
    if access_error:
        return access_error
    
    result = item.to_dict()
    result['required_materials'] = [rm.to_dict() for rm in item.required_materials]
    
    return jsonify(result), 200


@workplan_bp.route('/api/work-plan-items/<int:item_id>', methods=['PUT'])
@token_required
@role_required('client')
def update_work_plan_item(item_id):
    """Обновляет работу плана."""
    current_user = request.current_user
    
    item = db.get_or_404(WorkPlanItem, item_id)
    work_plan = db.get_or_404(WorkPlan, item.work_plan_id)
    
    access_error = require_project_access(work_plan.project_id, current_user['id'], current_user['role'])
    if access_error:
        return access_error
    
    data = request.get_json()
    if not data:
        return jsonify({'message': 'Отсутствуют данные для обновления'}), 400
    
    if 'name' in data:
        item.name = data['name']
    if 'quantity' in data:
        item.quantity = float(data['quantity'])
    if 'unit' in data:
        item.unit = data['unit']
    if 'status' in data:
        item.status = data['status']
    if 'progress' in data:
        progress = float(data['progress'])
        if 0 <= progress <= 100:
            item.progress = progress
    
    if 'start_date' in data and 'end_date' in data:
        try:
            item_start = datetime.fromisoformat(data['start_date'].replace('Z', '+00:00')).date()
            item_end = datetime.fromisoformat(data['end_date'].replace('Z', '+00:00')).date()
            
            if item_start >= item_end:
                return jsonify({'message': 'Дата начала должна быть раньше даты окончания'}), 400
            
            item.start_date = item_start
            item.end_date = item_end
        except ValueError:
            return jsonify({'message': 'Неверный формат даты'}), 400
    
    db.session.commit()
    
    result = item.to_dict()
    result['required_materials'] = [rm.to_dict() for rm in item.required_materials]
    
    return jsonify(result), 200


@workplan_bp.route('/api/work-plan-items/<int:item_id>/materials', methods=['POST'])
@token_required
@role_required('client')
def add_required_material(item_id):
    """Добавляет плановый материал к работе."""
    current_user = request.current_user
    
    item = db.get_or_404(WorkPlanItem, item_id)
    work_plan = db.get_or_404(WorkPlan, item.work_plan_id)
    
    access_error = require_project_access(work_plan.project_id, current_user['id'], current_user['role'])
    if access_error:
        return access_error
    
    data = request.get_json()
    if not data or not data.get('material_id') or not data.get('planned_quantity'):
        return jsonify({'message': 'Требуются material_id и planned_quantity'}), 400
    
    material = db.get_or_404(Material, data['material_id'])
    
    existing = RequiredMaterial.query.filter_by(
        work_item_id=item_id,
        material_id=material.id
    ).first()
    
    if existing:
        return jsonify({'message': 'Этот материал уже добавлен к работе'}), 409
    
    required_material = RequiredMaterial(
        work_item_id=item_id,
        material_id=material.id,
        planned_quantity=float(data['planned_quantity'])
    )
    db.session.add(required_material)
    db.session.commit()
    
    return jsonify(required_material.to_dict()), 201


@workplan_bp.route('/api/work-plan-items/<int:item_id>/materials/<int:material_id>', methods=['PUT'])
@token_required
@role_required('client')
def update_required_material(item_id, material_id):
    """Обновляет плановое количество материала."""
    current_user = request.current_user
    
    item = db.get_or_404(WorkPlanItem, item_id)
    work_plan = db.get_or_404(WorkPlan, item.work_plan_id)
    
    access_error = require_project_access(work_plan.project_id, current_user['id'], current_user['role'])
    if access_error:
        return access_error
    
    required_material = RequiredMaterial.query.filter_by(
        work_item_id=item_id,
        material_id=material_id
    ).first()
    
    if not required_material:
        return jsonify({'message': 'Плановый материал не найден'}), 404
    
    data = request.get_json()
    if not data or 'planned_quantity' not in data:
        return jsonify({'message': 'Требуется planned_quantity'}), 400
    
    required_material.planned_quantity = float(data['planned_quantity'])
    db.session.commit()
    
    return jsonify(required_material.to_dict()), 200


@workplan_bp.route('/api/work-plan-items/<int:item_id>/materials/<int:material_id>', methods=['DELETE'])
@token_required
@role_required('client')
def delete_required_material(item_id, material_id):
    """Удаляет плановый материал из работы."""
    current_user = request.current_user
    
    item = db.get_or_404(WorkPlanItem, item_id)
    work_plan = db.get_or_404(WorkPlan, item.work_plan_id)
    
    access_error = require_project_access(work_plan.project_id, current_user['id'], current_user['role'])
    if access_error:
        return access_error
    
    required_material = RequiredMaterial.query.filter_by(
        work_item_id=item_id,
        material_id=material_id
    ).first()
    
    if not required_material:
        return jsonify({'message': 'Плановый материал не найден'}), 404
    
    db.session.delete(required_material)
    db.session.commit()
    
    return jsonify({'message': 'Плановый материал успешно удален'}), 200


@workplan_bp.route('/api/materials', methods=['GET'])
@token_required
def get_materials():
    """Получает список всех материалов."""
    materials = Material.query.all()
    return jsonify([m.to_dict() for m in materials]), 200


@workplan_bp.route('/api/materials', methods=['POST'])
@token_required
@role_required('client')
def create_material():
    """Создает новый материал в справочнике."""
    data = request.get_json()
    if not data or not data.get('name') or not data.get('unit'):
        return jsonify({'message': 'Требуются name и unit'}), 400
    
    existing = Material.query.filter_by(name=data['name']).first()
    if existing:
        return jsonify({'message': 'Материал с таким названием уже существует'}), 409
    
    material = Material(
        name=data['name'],
        unit=data['unit']
    )
    db.session.add(material)
    db.session.commit()
    
    return jsonify(material.to_dict()), 201
